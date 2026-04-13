"""Rutas para métricas y exportación de reportes."""

import io
import os
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from database import get_db
from models import MetricasResponse
import openpyxl

AUDIOS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "audios")

router = APIRouter(tags=["Reportes"])


@router.get("/api/metricas", response_model=MetricasResponse)
async def obtener_metricas():
    """Retorna métricas generales del sistema de cobranza."""
    db = await get_db()
    try:
        cursor = await db.execute("SELECT COUNT(*) as total FROM deudores")
        total_deudores = (await cursor.fetchone())["total"]

        cursor = await db.execute(
            "SELECT COUNT(*) as total FROM deudores WHERE estado IN ('llamado', 'contesto', 'no_contesto')"
        )
        total_llamados = (await cursor.fetchone())["total"]

        cursor = await db.execute(
            "SELECT COUNT(*) as total FROM deudores WHERE estado = 'contesto'"
        )
        total_contestaron = (await cursor.fetchone())["total"]

        tasa = (total_contestaron / total_llamados * 100) if total_llamados > 0 else 0.0

        cursor = await db.execute("SELECT COALESCE(SUM(monto_deuda), 0) as total FROM deudores")
        monto_total = (await cursor.fetchone())["total"]

        return MetricasResponse(
            total_deudores=total_deudores,
            total_llamados=total_llamados,
            tasa_contestacion=round(tasa, 1),
            monto_total_mora=monto_total,
        )
    finally:
        await db.close()


@router.get("/api/audio/{deudor_id}")
async def servir_audio(deudor_id: int):
    """Sirve el archivo MP3 para reproducción."""
    audio_path = os.path.join(AUDIOS_DIR, f"cobranza_{deudor_id}.mp3")
    if not os.path.exists(audio_path):
        raise HTTPException(status_code=404, detail="Audio no encontrado")
    return FileResponse(audio_path, media_type="audio/mpeg", filename=f"cobranza_{deudor_id}.mp3")


@router.get("/api/exportar/excel")
async def exportar_excel():
    """Exporta reporte de deudores a Excel."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT nombre, rut, telefono, monto_deuda, dias_mora, email, estado, created_at FROM deudores ORDER BY dias_mora DESC"
        )
        rows = await cursor.fetchall()
    finally:
        await db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Cobranza"

    encabezados = ["Nombre", "RUT", "Teléfono", "Monto Deuda", "Días Mora", "Email", "Estado", "Fecha Registro"]
    ws.append(encabezados)

    for col_num, _ in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    for row in rows:
        r = dict(row)
        ws.append([
            r["nombre"], r["rut"], r["telefono"],
            r["monto_deuda"], r["dias_mora"],
            r["email"] or "", r["estado"], r["created_at"],
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_cobranza.xlsx"},
    )
